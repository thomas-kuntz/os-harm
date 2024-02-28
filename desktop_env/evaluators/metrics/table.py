import functools
import itertools
import logging
import os.path
# import operator
from numbers import Number
from typing import Any, Union, cast, Callable, Iterable
from typing import Dict, List, Tuple

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
# from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .utils import _match_value_to_rule, _read_cell_style, read_cell_value
from .utils import load_charts, load_sparklines, load_rows_or_cols, load_xlsx_styles\
                 , load_filters, load_pivot_tables

# from openpyxl.utils import coordinate_to_tuple

logger = logging.getLogger("desktopenv.metric.table")

BOOK = Union[pd.ExcelFile, Workbook, str]


def _parse_sheet_idx(sheet_idx: Union[int, str]
                     , result: BOOK, expected: BOOK
                     , result_sheet_names: List[str]
                     , expected_sheet_names: List[str]
                     ) -> Tuple[BOOK, str]:
    #  function _parse_sheet_idx {{{ # 
    if isinstance(sheet_idx, int):
        try:
            index: str = result_sheet_names[sheet_idx]
        except:
            index = ""
        book: BOOK = result
    elif sheet_idx.startswith("RI"):
        try:
            index: str = result_sheet_names[int(sheet_idx[2:])]
        except:
            index = ""
        book: BOOK = result
    elif sheet_idx.startswith("RN"):
        index: str = sheet_idx[2:]
        book: BOOK = result
    elif sheet_idx.startswith("EI"):
        try:
            index: str = expected_sheet_names[int(sheet_idx[2:])]
        except:
            index = ""
        book: BOOK = expected
    elif sheet_idx.startswith("EN"):
        index: str = sheet_idx[2:]
        book: BOOK = expected
    else:
        logger.error("Unrecognized sheet index")
        raise ValueError("Unrecognized sheet index")
    return book, index
    #  }}} function _parse_sheet_idx # 


SHEET = Union[pd.DataFrame, Worksheet, List[str]]


def _load_sheet(book: BOOK, index: str) -> SHEET:
    #  function _load_sheet {{{ # 
    try:
        if isinstance(book, str):
            book: str = cast(str, book)
            csv_name: str = "{:}-{:}.csv".format(os.path.splitext(book)[0], index)

            with open(csv_name) as f:
                csv_lines: List[str] = list(itertools.dropwhile(lambda l: len(l) == 0
                                                                , map(lambda l: l.strip()
                                                                      , reversed(f.read().splitlines())
                                                                      )
                                                                )
                                            )
            return csv_lines
        if isinstance(book, pd.ExcelFile):
            return pd.read_excel(book, index)
        if isinstance(book, Workbook):
            return book[index]
        logger.error("Not supported workbook format")
        raise NotImplementedError("Not supported workbook format")
    except NotImplementedError as e:
        raise e
    except:
        return None
    #  }}} function _load_sheet # 


def compare_table(result: str, expected: str = None, **options) -> float:
    #  function compare_table {{{ # 
    """
    Args:
        result (str): path to result xlsx
        expected (str): path to golden xlsx
        rules (List[Dict[str, Any]]): list of dict like
          {
            "type": str,
            <str as parameters>: anything
          }
          as sequential rules

    Returns:
        float: the score
    """

    if result is None:
        return 0.

    try:
        xlworkbookr: Workbook = openpyxl.load_workbook(filename=result)
        pdworkbookr = pd.ExcelFile(result)
    except:
        return 0.
    worksheetr_names: List[str] = pdworkbookr.sheet_names

    if expected is not None:

        xlworkbooke: Workbook = openpyxl.load_workbook(filename=expected)
        pdworkbooke = pd.ExcelFile(expected)
        worksheete_names: List[str] = pdworkbooke.sheet_names
    else:
        xlworkbooke: Workbook = None
        pdworkbooke = None
        worksheete_names: List[str] = None

    parse_idx: Callable[[Union[str, int], BOOK, BOOK], Tuple[BOOK, str]] = \
        functools.partial(
            _parse_sheet_idx,
            result_sheet_names=worksheetr_names,
            expected_sheet_names=worksheete_names
        )

    passes = True
    for r in options["rules"]:
        if r["type"] == "sheet_name":
            #  Compare Sheet Names {{{ # 
            metric: bool = worksheetr_names == worksheete_names
            logger.debug("Assertion: %s.sheet_names == %s.sheet_names - %s", result, expected, metric)
            #  }}} Compare Sheet Names # 

        elif r["type"] == "sheet_data":
            #  Compare Sheet Data by Internal Value {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # precision: int as number of decimal digits, default to 4

            error_limit: int = r.get("precision", 4)
            sheet1: pd.DataFrame = _load_sheet(*parse_idx(r["sheet_idx0"], pdworkbookr, pdworkbooke))
            if sheet1 is None:
                return 0.
            sheet2: pd.DataFrame = _load_sheet(*parse_idx(r["sheet_idx1"], pdworkbookr, pdworkbooke))

            sheet1 = sheet1.round()
            sheet2 = sheet2.round()
            metric: bool = sheet1.equals(sheet2)
            logger.debug("Sheet1: \n%s", str(sheet1))
            logger.debug("Sheet2: \n%s", str(sheet2))
            try:
                logger.debug("Sheet1 =v= Sheet2: \n%s", str(sheet1==sheet2))
            except:
                logger.debug("Sheet1 =/v= Sheet2")
            logger.debug("Assertion: %s =v= %s - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Compare Sheet Data by Internal Value # 

        elif r["type"] == "sheet_print":
            #  Compare Sheet Data by Printed Value {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # ignore_case: optional, defaults to False

            sheet1: List[str] = _load_sheet(*parse_idx(r["sheet_idx0"], result, expected))
            if sheet1 is None:
                return 0.
            sheet2: List[str] = _load_sheet(*parse_idx(r["sheet_idx1"], result, expected))
            if r.get("ignore_case", False):
                sheet1 = [l.lower() for l in sheet1]
                sheet2 = [l.lower() for l in sheet2]
            metric: bool = sheet1 == sheet2
            logger.debug("Assertion: %s =p= %s - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Compare Sheet Data by Printed Value # 

        elif r["type"] == "sparkline":
            #  Compare Sparklines {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            sparkline1: Dict[str, str] = load_sparklines(*parse_idx(r["sheet_idx0"], result, expected))
            sparkline2: Dict[str, str] = load_sparklines(*parse_idx(r["sheet_idx1"], result, expected))
            metric: bool = sparkline1 == sparkline2
            logger.debug("Assertion: %s.sp == %.sp - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Compare Sparklines # 

        elif r["type"] == "chart":
            #  Compare Charts {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # chart_props: list of str, see utils.load_charts

            charts1: Dict[str, Any] = load_charts(*parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r)
            charts2: Dict[str, Any] = load_charts(*parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r)
            metric: bool = charts1 == charts2
            logger.debug("Assertion: %s[chart] == %s[chart] - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Compare Charts # 

        elif r["type"] == "style":
            #  Compare Style (Also Conditional Formatiing) {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str indicating concerned styles, see utils._read_cell_style

            sheet_idx1: Tuple[BOOK, str] = parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke)
            book_name1: str = parse_idx(r["sheet_idx0"], result, expected)[0]
            styles1: Dict[str, List[Any]] = load_xlsx_styles(*sheet_idx1, book_name1, **r)

            sheet_idx2: Tuple[BOOK, str] = parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke)
            book_name2: str = parse_idx(r["sheet_idx1"], result, expected)[0]
            styles2: Dict[str, List[Any]] = load_xlsx_styles(*sheet_idx2, book_name2, **r)
            # number_formats1: List[str] = [c.number_format.lower() for col in sheet1.iter_cols() for c in col if c.value is not None and c.data_type=="n"]
            # number_formats2: List[str] = [c.number_format.lower() for col in sheet2.iter_cols() for c in col if c.value is not None and c.data_type=="n"]
            metric: bool = styles1 == styles2
            logger.debug("Assertion: %s.style == %s.style - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Compare Style (Also Conditional Formatiing) # 

        elif r["type"] == "freeze":
            #  Compare Freezing {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            sheet1: Worksheet = _load_sheet(*parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke))
            if sheet1 is None:
                return 0.
            sheet2: Worksheet = _load_sheet(*parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke))
            metric: bool = sheet1.freeze_panes == sheet2.freeze_panes
            logger.debug("Assertion: %s.freeze(%s) == %s.freeze(%s) - %s"
                         , r["sheet_idx0"], sheet1.freeze_panes
                         , r["sheet_idx1"], sheet2.freeze_panes
                         , metric
                         )
            #  }}} Compare Freezing # 

        elif r["type"] == "zoom":
            #  Check Zooming {{{ # 
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # method: str
            # ref: value

            sheet: Worksheet = _load_sheet(*parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke))
            if sheet is None:
                return 0.
            zoom_scale: Number = sheet.sheet_view.zoomScale or 100.
            metric: bool = _match_value_to_rule(zoom_scale, r)
            logger.debug("Assertion: %s.zoom(%.1f) %s %.1f - %s", r["sheet_idx"], zoom_scale, r["method"], r["ref"],
                         metric)
            #  }}} Check Zooming # 

        elif r["type"] == "data_validation":
            #  Check Data Validation {{{ # 
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # dv_props: list of dict like {attribute: {"method": str, "ref": anything}}
            #   available attributes:
            #     * ranges
            #     * type
            #     * formula1
            #     * formula2
            #     * operator
            #     * allowBlank
            #     * showDropDown
            #     * showInputMessage
            #     * showErrorMessage
            #     * error
            #     * errorTitle
            #     * errorStyle
            #     * prompt
            #     * promptTitle
            #     * imeMode

            sheet: Worksheet = _load_sheet(*parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke))
            if sheet is None:
                return 0.
            data_validators: List[DataValidation] = sheet.data_validations.dataValidation

            total_metric = len(data_validators) >= len(r["dv_props"])
            for dat_vldt in data_validators:
                metric = False
                for prpt in r["dv_props"]:
                    metric = metric or all(_match_value_to_rule(getattr(dat_vldt, attrbt)
                                                                , mr
                                                                ) \
                                           for attrbt, mr in prpt.items()
                                           )
                    if metric:
                        break
                total_metric = total_metric and metric
                if not total_metric:
                    break

            logger.debug("Assertion: %s.data_validation - %s", r["sheet_idx"], total_metric)
            metric: bool = total_metric
            #  }}} Check Data Validation # 

        elif r["type"] == "row_props":
            #  Check Row Properties {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str, see utils.load_rows_or_cols

            rows1: Dict[str, Any] = load_rows_or_cols(*parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke)
                                                      , obj="row"
                                                      , **r
                                                      )
            rows2: Dict[str, Any] = load_rows_or_cols(*parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke)
                                                      , obj="row"
                                                      , **r
                                                      )
            logger.debug("Rows1: %s", repr(rows1))
            logger.debug("Rows2: %s", repr(rows2))
            metric: bool = rows1 == rows2
            logger.debug("Assertion: %s[rows] == %s[rows] - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Check Row Properties # 

        elif r["type"] == "col_props":
            #  Check Row Properties {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str, see utils.load_rows_or_cols

            cols1: Dict[str, Any] = load_rows_or_cols(*parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke)
                                                      , obj="column"
                                                      , **r
                                                      )
            cols2: Dict[str, Any] = load_rows_or_cols(*parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke)
                                                      , obj="column"
                                                      , **r
                                                      )
            metric: bool = cols1 == cols2
            logger.debug("Assertion: %s[cols] == %s[cols] - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Check Row Properties # 

        elif r["type"] == "filter":
            #  Compare Filters {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            filters1: Dict[str, Any] = load_filters(*parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r)
            filters2: Dict[str, Any] = load_filters(*parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r)
            metric: bool = filters1==filters2
            logger.debug("Assertion: %s[filter] == %s[filter] - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Compare Filters # 

        elif r["type"] == "pivot_table":
            #  Compare Pivot Tables {{{ # 
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # pivot_props: list of str, see utils.load_pivot_tables

            pivots1: Dict[str, Any] = load_pivot_tables(*parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r)
            pivots2: Dict[str, Any] = load_pivot_tables(*parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r)
            metric: bool = pivots1==pivots2
            logger.debug("Assertion: %s[pivot]==%s[pivot] - %s", r["sheet_idx0"], r["sheet_idx1"], metric)
            #  }}} Compare Pivot Tables # 

        elif r["type"] == "check_cell":
            #  Check Cell Properties {{{ # 
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # coordinate: str, "E3"
            # props: dict like {attribute: {"method": str, "ref": anything}}
            #   supported attributes: value & those supported by utils._read_cell_style

            sheet: Worksheet = _load_sheet(*parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke))
            if sheet is None:
                return 0.
            # data_frame: pd.DataFrame = _load_sheet(*parse_idx(r["sheet_idx"], pdworkbookr, pdworkbooke))
            cell: Cell = sheet[r["coordinate"]]
            metric: bool = True
            for prpt, rule in r["props"].items():
                if prpt == "value":
                    val = read_cell_value(*parse_idx(r["sheet_idx"], result, expected), r["coordinate"])
                else:
                    val = _read_cell_style(prpt, cell)

                metric = metric and _match_value_to_rule(val, rule)

            logger.debug("Assertion: %s[%s] :%s - %s"
                         , r["sheet_idx"], r["coordinate"]
                         , repr(r["props"]), metric
                         )
            #  }}} Check Cell Properties # 

        else:
            raise NotImplementedError("Unimplemented sheet check: {:}".format(r["type"]))

        passes = passes and metric
        if not passes:
            break

    return float(passes)
    #  }}} function compare_table # 


def compare_csv(result: str, expected: str, **options) -> float:
    if result is None:
        return 0.

    with open(result) as f:
        result_lines: Iterable[str] = f.read().splitlines()
    with open(expected) as f:
        expected_lines: Iterable[str] = f.read().splitlines()
    if not options.get("strict", True):
        result_lines = map(str.strip, result_lines)
        expected_lines = map(str.strip, expected_lines)
    if options.get("ignore_case", False):
        result_lines = map(str.lower, result_lines)
        expected_lines = map(str.lower, expected_lines)

    metric: bool = list(result_lines) == list(expected_lines)
    return float(metric)


if __name__ == '__main__':
    import datetime
    import sys

    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    datetime_str: str = datetime.datetime.now().strftime("%Y%m%d@%H%M%S")

    file_handler = logging.FileHandler(os.path.join("logs", "normal-{:}.log".format(datetime_str)))
    debug_handler = logging.FileHandler(os.path.join("logs", "debug-{:}.log".format(datetime_str)))
    stdout_handler = logging.StreamHandler(sys.stdout)
    sdebug_handler = logging.FileHandler(os.path.join("logs", "sdebug-{:}.log".format(datetime_str)))

    file_handler.setLevel(logging.INFO)
    debug_handler.setLevel(logging.DEBUG)
    stdout_handler.setLevel(logging.INFO)
    sdebug_handler.setLevel(logging.DEBUG)

    formatter = logging.Formatter(
        fmt="\x1b[1;33m[%(asctime)s \x1b[31m%(levelname)s \x1b[32m%(module)s/%(lineno)d-%(processName)s\x1b[1;33m] \x1b[0m%(message)s")
    file_handler.setFormatter(formatter)
    debug_handler.setFormatter(formatter)
    stdout_handler.setFormatter(formatter)
    sdebug_handler.setFormatter(formatter)

    stdout_handler.addFilter(logging.Filter("desktopenv"))
    sdebug_handler.addFilter(logging.Filter("desktopenv"))

    logger.addHandler(file_handler)
    logger.addHandler(debug_handler)
    logger.addHandler(stdout_handler)
    logger.addHandler(sdebug_handler)

    path1 = "snapshots/test/cache/4e6fcf72-daf3-439f-a232-c434ce416af6/Employee_Age_By_Birthday.xlsx"
    path2 = "snapshots/test/cache/4e6fcf72-daf3-439f-a232-c434ce416af6/Employee_Age_By_Birthday_gold.xlsx"
    rules = [ { "type": "sheet_data"
              , "sheet_idx0": 0
              , "sheet_idx1": "EI0"
              }
            ]
    print(compare_table(path1, path2
                        , rules=rules
                        )
          )
    print(compare_table(path2, path2
                        , rules=rules
                        )
          )

    # Row Properties
    # path1 = "../../任务数据/LibreOffice Calc/Date_Budget_Variance_HideNA.xlsx"
    # path2 = "../../任务数据/LibreOffice Calc/Date_Budget_Variance_HideNA_gold.xlsx"
    # workbook: Workbook = openpyxl.load_workbook(filename=path1)
    # worksheet: Worksheet = workbook.active
    # for r_no, dms in worksheet.column_dimensions.items():
    # print(r_no, type(r_no), type(dms), dms.hidden)

    # Conditional Formats
    # import formulas
    # path1 = "../../任务数据/LibreOffice Calc/Calendar_Highlight_Weekend_Days.xlsx"
    # path2 = "../../任务数据/LibreOffice Calc/Calendar_Highlight_Weekend_Days_gold.xlsx"
    # path3 = "../../任务数据/LibreOffice Calc/Calendar_Highlight_Weekend_Days_gold_test.xlsx"
    # workbook: Workbook = openpyxl.load_workbook(filename=path2)
    # worksheet: Worksheet = workbook.active
    # print(worksheet.conditional_formatting)
    # for itm in worksheet.conditional_formatting:
    # print(itm.cells)
    # for r in itm.rules:
    # print( r.type, r.formula, r.dxf.font.color.rgb
    # , r.dxf.fill.fgColor.rgb, r.dxf.fill.bgColor.rgb
    # )
    # condition = formulas.Parser().ast("=" + r.formula[0])[1].compile()
    ##print(r.type, r.operator, r.dxfId, r.dxf)
    # for r in itm.cells:
    # for c in r.cells:
    # value = worksheet.cell(row=c[0], column=c[1]).value
    # print(value, condition(str(value)))
